"""Extract plain text from uploaded files (PDF, Excel, CSV/TSV, text)."""
from __future__ import annotations

import os

SUPPORTED = {".pdf", ".xlsx", ".xlsm", ".csv", ".tsv", ".txt", ".md", ".json", ".log"}


def extract_text(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return _pdf(path)
    if ext in {".xlsx", ".xlsm"}:
        return _xlsx(path)
    if ext in {".csv", ".tsv", ".txt", ".md", ".json", ".log"}:
        with open(path, encoding="utf-8", errors="replace") as fh:
            return fh.read()
    raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(sorted(SUPPORTED))}")


def _pdf(path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    return "\n".join(pages)


def _xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"### Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)
