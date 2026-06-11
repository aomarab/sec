"""Export a report's Markdown tables to CSV or Excel. Works for any report
(briefing, analysis, scan) by extracting the GitHub-flavoured tables from the
saved .md file — so a single code path covers every report type."""
from __future__ import annotations

import csv
import io
import re

_TAG = re.compile(r"<[^>]+>")
_SEP = re.compile(r"^\s*\|?[\s:|\-]+\|?\s*$")


def _clean(cell: str) -> str:
    cell = _TAG.sub("", cell)                       # strip HTML (links, badges)
    cell = re.sub(r"\*\*(.+?)\*\*", r"\1", cell)    # bold
    cell = cell.replace("`", "").replace("\\|", "|")
    return cell.strip()


def extract_tables(markdown: str) -> list[dict]:
    lines = markdown.splitlines()
    tables: list[dict] = []
    heading = "Report"
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        hm = re.match(r"^#{1,6}\s+(.*)", line)
        if hm:
            heading = _clean(hm.group(1))
            i += 1
            continue
        if "|" in line and i + 1 < len(lines) and _SEP.match(lines[i + 1]):
            headers = [_clean(c) for c in line.strip().strip("|").split("|")]
            i += 2
            rows = []
            while i < len(lines) and "|" in lines[i] and lines[i].strip():
                rows.append([_clean(c) for c in lines[i].strip().strip("|").split("|")])
                i += 1
            tables.append({"title": heading, "headers": headers, "rows": rows})
            continue
        i += 1
    return tables


def to_csv(tables: list[dict]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    if not tables:
        w.writerow(["No tabular data in this report."])
        return buf.getvalue()
    for t in tables:
        w.writerow([t["title"]])
        w.writerow(t["headers"])
        for row in t["rows"]:
            w.writerow(row)
        w.writerow([])
    return buf.getvalue()


def _sheet_name(title: str, idx: int, used: set) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "", title)[:28].strip() or f"Table {idx}"
    base, n = name, 2
    while name.lower() in used:
        name = f"{base[:25]} {n}"
        n += 1
    used.add(name.lower())
    return name


def to_xlsx_bytes(tables: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    used: set = set()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F2A43")

    for idx, t in enumerate(tables, 1):
        ws = wb.create_sheet(_sheet_name(t["title"], idx, used))
        ws.append(t["headers"])
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
        for row in t["rows"]:
            ws.append(row)
        for col in range(1, len(t["headers"]) + 1):
            width = max([len(str(t["headers"][col - 1]))] +
                        [len(str(r[col - 1])) for r in t["rows"] if col - 1 < len(r)] + [8])
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(width + 2, 60)
    if not tables:
        wb.create_sheet("Report").append(["No tabular data in this report."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
